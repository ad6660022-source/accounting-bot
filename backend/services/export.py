"""
Генерация Excel-выгрузки по воркспейсу.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

from backend.database.models import TX_LABELS, TxType

# ── Группы операций ───────────────────────────────────────────────────────────
_INCOME_TYPES   = {TxType.PRIHOD_MES, TxType.PRIHOD_FAST, TxType.PRIHOD_STO}
_EXPENSE_TYPES  = {TxType.ZAKUP, TxType.STORONNIE, TxType.EXPENSE_WRITEOFF}
_DEBT_TYPES     = {TxType.ODOLZHIT, TxType.POGASIT}
_TRANSFER_TYPES = {TxType.SNYAT_RS, TxType.SNYAT_DEBIT, TxType.VNESTI_RS}

# ── Стили ─────────────────────────────────────────────────────────────────────
_H_FILL   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_H_FONT   = Font(color="FFFFFF", bold=True, size=11)
_H_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_C_ALIGN  = Alignment(vertical="center")
_R_ALIGN  = Alignment(horizontal="right", vertical="center")

_GREEN_FILL  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_RED_FILL    = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
_GRAY_FILL   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
_BLUE_FILL   = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

_GREEN_FONT  = Font(color="2E7D32", bold=True)
_RED_FONT    = Font(color="C62828", bold=True)
_BLUE_FONT   = Font(color="1565C0", bold=True)

_THIN_BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)


def _style_header(ws, headers: list[str], row_height: int = 36):
    ws.append(headers)
    ws.row_dimensions[1].height = row_height
    for cell in ws[1]:
        cell.font = _H_FONT
        cell.fill = _H_FILL
        cell.alignment = _H_ALIGN


def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt(v: int) -> str:
    return f"{v:,}".replace(",", " ") if v else "0"


def _fmt_delta(v: int) -> str:
    if v == 0: return ""
    sign = "+" if v > 0 else ""
    return f"{sign}{v:,}".replace(",", " ")


def _get_delta(tx) -> tuple[int, int, int]:
    t, a = tx.type, tx.amount
    dest = tx.destination or "cash"
    if t in (TxType.ZAKUP, TxType.STORONNIE):
        return (-a, 0, 0)
    elif t == TxType.EXPENSE_WRITEOFF:
        if dest == "bank":   return (0, -a, 0)
        elif dest == "debit": return (0, 0, -a)
        return (-a, 0, 0)
    elif t in _INCOME_TYPES:
        if dest == "bank":   return (0, a, 0)
        elif dest == "debit": return (0, 0, a)
        return (a, 0, 0)
    elif t == TxType.SNYAT_RS:    return (0, -a, a)
    elif t == TxType.SNYAT_DEBIT: return (a, 0, -a)
    elif t == TxType.VNESTI_RS:   return (-a, a, 0)
    elif t in (TxType.ODOLZHIT, TxType.POGASIT): return (-a, 0, 0)
    return (0, 0, 0)


def _compute_rows(ip, txs_asc: list) -> list[tuple]:
    cash, bank, debit = ip.cash_balance, ip.bank_balance, ip.debit_balance
    rows = []
    for tx in reversed(txs_asc):
        dc, db, dd = _get_delta(tx)
        rows.append((tx, cash, bank, debit, dc, db, dd))
        cash -= dc; bank -= db; debit -= dd
    rows.reverse()
    return rows


def _add_tx_sheet(ws, rows: list[tuple], type_filter: Optional[set] = None):
    headers = [
        "Дата", "Время", "ИП", "Тип операции", "Источник", "Кто провёл",
        "Комментарий", "Изм. Нал", "Изм. Р/С", "Изм. Дебет",
        "Баланс Нал", "Баланс Р/С", "Баланс Дебет",
    ]
    _style_header(ws, headers)
    _set_widths(ws, [12, 8, 18, 22, 9, 18, 28, 13, 13, 13, 15, 15, 15])

    r = 2
    for tx, bal_cash, bal_bank, bal_debit, dc, db, dd in rows:
        if tx.is_cancelled:
            continue
        if type_filter and tx.type not in type_filter:
            continue
        dt: datetime = tx.created_at
        dest_label = {"cash": "Нал", "bank": "Р/С", "debit": "Дебет"}.get(tx.destination or "cash", "")
        row_data = [
            dt.strftime("%d.%m.%Y"),
            dt.strftime("%H:%M"),
            tx.ip.name if tx.ip else "",
            TX_LABELS.get(tx.type, tx.type),
            dest_label,
            tx.user.display_name if tx.user else "",
            tx.comment or "",
            _fmt_delta(dc), _fmt_delta(db), _fmt_delta(dd),
            _fmt(bal_cash), _fmt(bal_bank), _fmt(bal_debit),
        ]
        ws.append(row_data)
        ws.row_dimensions[r].height = 20

        total = dc + db + dd
        fill = _GREEN_FILL if total > 0 else (_RED_FILL if total < 0 else None)
        if fill:
            for c in range(1, 14):
                ws.cell(r, c).fill = fill
        for c in range(1, 14):
            ws.cell(r, c).border = _THIN_BORDER
        for c in range(8, 14):
            ws.cell(r, c).alignment = _R_ALIGN
        r += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M1"


def _add_analytics_sheet(ws, ips: list, all_txs: list, date_from=None, date_to=None):
    """Лист сводной аналитики по всем ИП."""

    def in_range(tx) -> bool:
        d = tx.created_at.date()
        if date_from and d < date_from: return False
        if date_to   and d > date_to:   return False
        return True

    txs = [t for t in all_txs if not t.is_cancelled and in_range(t)]

    # --- Итоги по типам операций ---
    ws.append(["АНАЛИТИКА ПО ОПЕРАЦИЯМ"])
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.row_dimensions[1].height = 28
    ws.append([])

    by_type: dict[str, int] = {}
    for tx in txs:
        by_type[tx.type] = by_type.get(tx.type, 0) + tx.amount

    income_types = [
        (TxType.PRIHOD_FAST, "⚡ Приход быстрый"),
        (TxType.PRIHOD_MES,  "📥 Приход ежемесячный"),
        (TxType.PRIHOD_STO,  "🏦 Приход сторонний"),
    ]
    expense_types = [
        (TxType.ZAKUP,            "🛒 Закупы"),
        (TxType.STORONNIE,        "💸 Посторонние траты"),
        (TxType.EXPENSE_WRITEOFF, "💰 Расходы (списания)"),
    ]
    other_types = [
        (TxType.SNYAT_RS,    "Снятие с Р/С"),
        (TxType.SNYAT_DEBIT, "Снятие с дебета"),
        (TxType.VNESTI_RS,   "Внесение на Р/С"),
        (TxType.ODOLZHIT,    "Займ выдан"),
        (TxType.POGASIT,     "Займ погашен"),
    ]

    def section(title, rows_data, header_fill, val_font):
        ws.append([title])
        last = ws.max_row
        ws[f"A{last}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{last}"].fill = header_fill
        ws.row_dimensions[last].height = 22
        ws.merge_cells(f"A{last}:C{last}")

        hrow = ws.max_row + 1
        ws.append(["Тип операции", "Сумма", "Кол-во"])
        for c in ["A", "B", "C"]:
            ws[f"{c}{hrow}"].font = Font(bold=True, color="FFFFFF")
            ws[f"{c}{hrow}"].fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
            ws[f"{c}{hrow}"].alignment = _H_ALIGN
        ws.row_dimensions[hrow].height = 20

        total = 0
        for tx_type, label in rows_data:
            amt = by_type.get(tx_type, 0)
            cnt = sum(1 for t in txs if t.type == tx_type)
            dr = ws.max_row + 1
            ws.append([label, amt, cnt])
            ws[f"B{dr}"].number_format = '#,##0'
            ws[f"B{dr}"].alignment = _R_ALIGN
            ws[f"B{dr}"].font = val_font
            ws[f"C{dr}"].alignment = _R_ALIGN
            ws.row_dimensions[dr].height = 18
            for c in range(1, 4):
                ws.cell(dr, c).border = _THIN_BORDER
            total += amt

        tr = ws.max_row + 1
        ws.append(["ИТОГО", total, ""])
        ws[f"A{tr}"].font = Font(bold=True)
        ws[f"B{tr}"].font = val_font
        ws[f"B{tr}"].number_format = '#,##0'
        ws[f"B{tr}"].alignment = _R_ALIGN
        ws[f"B{tr}"].fill = _YELLOW_FILL
        ws[f"A{tr}"].fill = _YELLOW_FILL
        ws.row_dimensions[tr].height = 22
        ws.append([])

    section("ПОСТУПЛЕНИЯ",
            income_types,
            PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
            _GREEN_FONT)

    section("РАСХОДЫ",
            expense_types,
            PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
            _RED_FONT)

    section("ПРОЧИЕ ОПЕРАЦИИ",
            other_types,
            PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid"),
            _BLUE_FONT)

    # --- Итоги по ИП ---
    ws.append([])
    r = ws.max_row + 1
    ws.append(["ОСТАТКИ ПО ИП"])
    ws[f"A{r}"].font = Font(bold=True, size=14, color="1F4E79")
    ws.row_dimensions[r].height = 28

    hrow = ws.max_row + 1
    ws.append(["ИП", "Нал", "Р/С", "Дебет", "Итого"])
    for c in ["A", "B", "C", "D", "E"]:
        ws[f"{c}{hrow}"].font = Font(bold=True, color="FFFFFF")
        ws[f"{c}{hrow}"].fill = _H_FILL
        ws[f"{c}{hrow}"].alignment = _H_ALIGN
    ws.row_dimensions[hrow].height = 24

    for ip in ips:
        total = (ip.cash_balance or 0) + (ip.bank_balance or 0) + (ip.debit_balance or 0)
        dr = ws.max_row + 1
        ws.append([ip.name, ip.cash_balance or 0, ip.bank_balance or 0, ip.debit_balance or 0, total])
        for c in range(1, 6):
            ws.cell(dr, c).border = _THIN_BORDER
        for c in range(2, 6):
            ws.cell(dr, c).number_format = '#,##0'
            ws.cell(dr, c).alignment = _R_ALIGN
        ws.row_dimensions[dr].height = 20

    _set_widths(ws, [30, 18, 18, 18, 18])
    ws.freeze_panes = "A2"


def _add_expenses_sheet(ws, expenses: list):
    """Лист расходов со списаниями."""
    headers = ["Описание расхода", "Сумма заявл.", "Списано", "Остаток", "Дата создания", "Закрыт"]
    _style_header(ws, headers)
    _set_widths(ws, [35, 15, 15, 15, 16, 10])

    for exp in expenses:
        written = sum(w.amount for w in (exp.writeoffs or []) if not w.is_cancelled)
        remaining = exp.amount - written
        dr = ws.max_row + 1
        ws.append([
            exp.description,
            exp.amount, written, remaining,
            exp.created_at.strftime("%d.%m.%Y"),
            "Да" if exp.is_closed else "Нет",
        ])
        ws.row_dimensions[dr].height = 20
        for c in [2, 3, 4]:
            ws.cell(dr, c).number_format = '#,##0'
            ws.cell(dr, c).alignment = _R_ALIGN
        if exp.is_closed:
            ws.cell(dr, 1).font = Font(color="888888")
        fill = _GREEN_FILL if remaining == 0 else (_YELLOW_FILL if written > 0 else None)
        if fill:
            for c in range(1, 7):
                ws.cell(dr, c).fill = fill
        for c in range(1, 7):
            ws.cell(dr, c).border = _THIN_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"


def generate_excel(
    ip,
    all_transactions: list,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> bytes:
    """Выгрузка по одному ИП (legacy, используется старым endpoint)."""
    sorted_txs = sorted(all_transactions, key=lambda t: t.created_at)
    all_rows = _compute_rows(ip, sorted_txs)

    def in_range(tx) -> bool:
        d = tx.created_at.date()
        if date_from and d < date_from: return False
        if date_to   and d > date_to:   return False
        return True

    filtered = [(tx, c, b, d, dc, db, dd) for (tx, c, b, d, dc, db, dd) in all_rows if in_range(tx)]

    wb = Workbook()
    ws1 = wb.active; ws1.title = "Все операции"
    _add_tx_sheet(ws1, filtered)
    ws2 = wb.create_sheet("Приходы")
    _add_tx_sheet(ws2, filtered, _INCOME_TYPES)
    ws3 = wb.create_sheet("Расходы")
    _add_tx_sheet(ws3, filtered, _EXPENSE_TYPES)
    ws4 = wb.create_sheet("Займы")
    _add_tx_sheet(ws4, filtered, _DEBT_TYPES)
    ws5 = wb.create_sheet("Переводы")
    _add_tx_sheet(ws5, filtered, _TRANSFER_TYPES)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_workspace_excel(
    ips: list,
    all_transactions: list,
    expenses: list,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> bytes:
    """Полная выгрузка по воркспейсу — все ИП, аналитика, расходы."""
    sorted_txs = sorted(all_transactions, key=lambda t: t.created_at)

    # Running balance для каждого ИП отдельно
    all_rows: list[tuple] = []
    for ip in ips:
        ip_txs = [t for t in sorted_txs if t.ip_id == ip.id]
        all_rows.extend(_compute_rows(ip, ip_txs))

    # Сортируем объединённый список по дате
    all_rows.sort(key=lambda x: x[0].created_at)

    def in_range(tx) -> bool:
        d = tx.created_at.date()
        if date_from and d < date_from: return False
        if date_to   and d > date_to:   return False
        return True

    filtered = [(tx, c, b, d, dc, db, dd) for (tx, c, b, d, dc, db, dd) in all_rows if in_range(tx)]

    wb = Workbook()

    # Лист 1 — Аналитика
    ws_a = wb.active; ws_a.title = "Аналитика"
    txs_for_analytics = [t for t in sorted_txs if in_range(t)]
    _add_analytics_sheet(ws_a, ips, txs_for_analytics, date_from, date_to)

    # Лист 2 — Все операции
    ws1 = wb.create_sheet("Все операции")
    _add_tx_sheet(ws1, filtered)

    # Лист 3 — Приходы
    ws2 = wb.create_sheet("Приходы")
    _add_tx_sheet(ws2, filtered, _INCOME_TYPES)

    # Лист 4 — Расходы
    ws3 = wb.create_sheet("Расходы")
    _add_tx_sheet(ws3, filtered, _EXPENSE_TYPES)

    # Лист 5 — Переводы
    ws4 = wb.create_sheet("Переводы")
    _add_tx_sheet(ws4, filtered, _TRANSFER_TYPES)

    # Лист 6 — Займы
    ws5 = wb.create_sheet("Займы")
    _add_tx_sheet(ws5, filtered, _DEBT_TYPES)

    # Лист 7 — Расходы (статьи)
    ws6 = wb.create_sheet("Статьи расходов")
    _add_expenses_sheet(ws6, expenses)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
